# views.py
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib import colors
import os
import tempfile
import datetime
from django.views.decorators.csrf import csrf_protect
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.db import IntegrityError
from django.utils import timezone
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.db.models import Q, Count, Prefetch, Sum
from django.core.paginator import Paginator
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
from django.contrib.auth.models import User, Group
from django.core.exceptions import MultipleObjectsReturned, ValidationError

from .models import Trabajo, Maquina, Faena, Empresa, PerfilUsuario # Asegúrate que PerfilUsuario esté aquí
from .forms import (
    TrabajoForm, MaquinaForm, FaenaForm, UserRegistrationForm,
    UserEditForm, EmpresaForm, EmpresaEditForm
)
from .filters import TrabajoFilter # Asegúrate que este filtro exista

# ---------------------- HELPERS ----------------------
def get_user_empresa(user):
    """
    Obtiene la única empresa asociada a un usuario no superusuario.
    Prioriza 'empresas_administradas' y luego 'perfil.empresa'.
    Devuelve la instancia de Empresa si es única.
    Devuelve None si no hay empresa o es superusuario.
    Lanza MultipleObjectsReturned si se encuentran múltiples empresas administradas.
    """
    if user.is_superuser:
        return None

    # 1. Intenta obtener desde 'empresas_administradas' (prioridad)
    if hasattr(user, 'empresas_administradas'):
        empresas_admin = user.empresas_administradas.all()
        count = empresas_admin.count()

        if count == 1:
            # Caso ideal: administra exactamente una empresa
            return empresas_admin.first()
        elif count > 1:
            # Administra múltiples empresas: hay ambigüedad
            raise MultipleObjectsReturned(f"El usuario {user.username} administra {count} empresas.")
        # Si count es 0, continúa para verificar el perfil

    # 2. Si no se encontró una única en 'empresas_administradas', intenta desde el perfil
    if hasattr(user, 'perfil') and user.perfil and hasattr(user.perfil, 'empresa') and user.perfil.empresa:
        # Asume que la relación perfil-empresa implica una única empresa
        return user.perfil.empresa

    # 3. Si no se encontró empresa por ninguna vía
    return None

# ---------------------- VISTAS GENERALES ----------------------
@login_required
@csrf_protect
def custom_logout(request):
    if request.method == 'POST':
        logout(request)
        return redirect('login')
    return redirect('home')

def home(request):
    context = {}
    if request.user.is_authenticated:
        perfil_usuario = getattr(request.user, 'perfil', None)
        empresa = perfil_usuario.empresa if perfil_usuario else None
        context.update({
            'empresa': empresa,
            'es_administrador': request.user.groups.filter(name='Admin').exists(),
            'es_supervisor': request.user.groups.filter(name='Supervisor').exists()
        })
        if empresa and empresa.logo:
            context['logo_empresa'] = empresa.logo.url
    return render(request, 'registros/home.html', context)

def landing_page(request):
    return render(request, 'registros/landing_page.html')

# ---------------------- VISTAS DE EMPRESA ----------------------

@login_required
#@user_passes_test(lambda u: u.is_superuser) # O permiso específico
def crear_empresa(request):
    if not request.user.is_superuser:
         messages.error(request, "Solo los superusuarios pueden crear empresas.")
         return redirect('home') # O a donde corresponda

    if request.method == 'POST':
        # No se pasa 'user' aquí, EmpresaForm maneja la creación del admin
        form = EmpresaForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                empresa = form.save()
                messages.success(request, f'Empresa {empresa.nombre} creada exitosamente con su administrador!')
                return redirect('listar_empresas')
            except Exception as e:
                messages.error(request, f'Error al crear empresa: {str(e)}')
                # Considera añadir errores específicos del formulario si es necesario
                # Por ejemplo, si la creación del usuario falla dentro de form.save()
        else:
            # Mostrar errores de validación
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field if field != "__all__" else "General"}: {error}')
    else:
        form = EmpresaForm()
    return render(request, 'registros/crear_empresa.html', {'form': form})

@login_required
#@user_passes_test(lambda u: u.is_superuser)
def listar_empresas(request):
    if not request.user.is_superuser:
         messages.error(request, "Solo los superusuarios pueden listar empresas.")
         return redirect('home')

    empresas_list = Empresa.objects.all().order_by('nombre')
    paginator = Paginator(empresas_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'registros/listar_empresas.html', {'page_obj': page_obj})

@login_required
#@user_passes_test(lambda u: u.is_superuser) # Asegúrate que solo SU pueda editar
def editar_empresa(request, pk):
    if not request.user.is_superuser:
        messages.error(request, "Solo los superusuarios pueden editar empresas.")
        return redirect('home')

    empresa = get_object_or_404(Empresa, pk=pk)

    if request.method == 'POST':
        # Usa el nuevo formulario de edición
        form = EmpresaEditForm(request.POST, request.FILES, instance=empresa)
        if form.is_valid():
            form.save() # Guarda los cambios en la empresa (incluidos los límites)
            messages.success(request, f'Empresa "{empresa.nombre}" actualizada correctamente.')
            return redirect('listar_empresas') # Vuelve a la lista
        else:
            # Muestra errores de validación específicos
            for field, errors in form.errors.items():
                for error in errors:
                    # Muestra el error en la página de la lista, ya que el modal se cierra
                    messages.error(request, f'Error al actualizar {empresa.nombre} - {field if field != "__all__" else "General"}: {error}')
            # Redirige de vuelta a la lista para mostrar los mensajes de error
            # Idealmente, se podría usar AJAX para mantener el modal abierto y mostrar errores,
            # pero esta es la solución más simple sin AJAX.
            return redirect('listar_empresas')
    else:
        # Para GET (aunque no se usa si todo es por modal), preparamos el form
        # Si quisieras una página de edición separada, aquí usarías EmpresaEditForm
        # form = EmpresaEditForm(instance=empresa)
        # return render(request, 'registros/editar_empresa.html', {'form': form, 'empresa': empresa})
        # Como usamos modal, la petición GET no debería llegar a esta vista separada
        # Redirigir si se intenta acceder por GET podría ser una opción
        messages.warning(request, "La edición se realiza a través del modal en la lista.")
        return redirect('listar_empresas')
    
@login_required
@require_POST # Para consistencia con modal (si se implementa)
#@user_passes_test(lambda u: u.is_superuser)
def eliminar_empresa(request, pk):
    if not request.user.is_superuser:
         messages.error(request, "Solo los superusuarios pueden eliminar empresas.")
         return redirect('home')

    empresa = get_object_or_404(Empresa, pk=pk)
    try:
        empresa_nombre = empresa.nombre
        # Considerar qué pasa con usuarios, máquinas, etc., al eliminar la empresa
        # Podrías necesitar eliminar o desasociar objetos relacionados primero
        empresa.delete()
        messages.success(request, f'Empresa "{empresa_nombre}" eliminada correctamente')
    except Exception as e:
        messages.error(request, f'Error al eliminar empresa: {str(e)}')

    return redirect('listar_empresas')

@login_required
@user_passes_test(lambda u: u.is_superuser)
def detalles_empresa(request, pk):
    empresa = get_object_or_404(Empresa, pk=pk)

    # Conteo simple de usuarios de la empresa (sin importar el rol)
    # Esto asegura que se cuenten todos los usuarios correctamente
    usuarios_count = PerfilUsuario.objects.filter(empresa=empresa).count()
    
    # Contar máquinas
    maquina_count = Maquina.objects.filter(empresa=empresa).count()
    
    # Contar faenas
    faena_count = Faena.objects.filter(empresa=empresa).count()
    
    # Obtener faenas activas
    faenas_activas = Faena.objects.filter(
        empresa=empresa, 
        estado='activa'
    ).order_by('-fecha_inicio')[:5]
    
    # Contar trabajos
    trabajo_count = Trabajo.objects.filter(empresa=empresa).count()
    
    # Obtener trabajos por mes para el gráfico de evolución (últimos 6 meses)
    from django.db.models.functions import TruncMonth
    from datetime import datetime, timedelta
    
    # Definir rango de fechas (últimos 6 meses)
    fecha_fin = datetime.now().date() + timedelta(days=1)  # Incluir hoy
    fecha_inicio = (datetime.now() - timedelta(days=180)).date()  # 6 meses atrás
    
    # Obtener conteo de trabajos por mes
    trabajos_por_mes = Trabajo.objects.filter(
        empresa=empresa,
        fecha__gte=fecha_inicio,
        fecha__lt=fecha_fin
    ).annotate(
        mes=TruncMonth('fecha')
    ).values('mes').annotate(
        cantidad=Count('id')
    ).order_by('mes')
    
    # Convertir a un formato más fácil de usar en el template
    etiquetas_meses = []
    datos_trabajos = []
    
    # Crear mapa de meses para llenar meses sin datos
    meses_map = {}
    current_date = fecha_inicio.replace(day=1)
    while current_date < fecha_fin:
        month_name = current_date.strftime("%b %Y")
        meses_map[month_name] = 0
        etiquetas_meses.append(month_name)
        current_date = (current_date.replace(day=28) + timedelta(days=4)).replace(day=1)
    
    # Llenar datos reales
    for item in trabajos_por_mes:
        month_name = item['mes'].strftime("%b %Y")
        meses_map[month_name] = item['cantidad']
    
    # Convertir mapa a lista ordenada
    for month in etiquetas_meses:
        datos_trabajos.append(meses_map.get(month, 0))
    
    # Obtener las máquinas más utilizadas (top 5)
    maquinas_mas_usadas = Trabajo.objects.filter(
        empresa=empresa
    ).values(
        'maquina__nombre'
    ).annotate(
        cantidad=Count('id')
    ).order_by('-cantidad')[:5]
    
    etiquetas_maquinas = []
    datos_maquinas = []
    
    for item in maquinas_mas_usadas:
        etiquetas_maquinas.append(item['maquina__nombre'])
        datos_maquinas.append(item['cantidad'])
    
    # Obtener distribución de usuarios por rol de manera simple
    # Evitamos consultas complejas que pueden no funcionar
    user_roles = []
    
    # Contar usuarios administradores
    admin_count = User.objects.filter(
        groups__name='Admin',
        perfil__empresa=empresa
    ).count()
    
    if admin_count > 0:
        user_roles.append({
            'groups__name': 'Admin',
            'user_count': admin_count
        })
    
    # Contar usuarios supervisores
    supervisor_count = User.objects.filter(
        groups__name='Supervisor',
        perfil__empresa=empresa
    ).count()
    
    if supervisor_count > 0:
        user_roles.append({
            'groups__name': 'Supervisor',
            'user_count': supervisor_count
        })
    
    # Contar usuarios trabajadores
    trabajador_count = User.objects.filter(
        groups__name='Trabajador',
        perfil__empresa=empresa
    ).count()
    
    if trabajador_count > 0:
        user_roles.append({
            'groups__name': 'Trabajador',
            'user_count': trabajador_count
        })
    
    # Si el conteo de roles no suma el total, añadir "Sin rol asignado"
    total_roles = admin_count + supervisor_count + trabajador_count
    if total_roles < usuarios_count and (usuarios_count - total_roles) > 0:
        user_roles.append({
            'groups__name': 'Sin Rol Asignado',
            'user_count': usuarios_count - total_roles
        })
    
    # Obtener estados de trabajos para la gráfica de estados
    estados_trabajo = {
        'pendiente': Trabajo.objects.filter(empresa=empresa, estado='pendiente').count(),
        'aprobado': Trabajo.objects.filter(empresa=empresa, estado='aprobado').count(),
        'rechazado': Trabajo.objects.filter(empresa=empresa, estado='rechazado').count()
    }
    
    # Eliminar estados con 0 trabajos
    trabajos_por_estado = {k: v for k, v in estados_trabajo.items() if v > 0}
    
    # Crear contexto con TODOS los datos necesarios
    context = {
        'empresa': empresa,
        'usuarios_count': usuarios_count,
        'user_roles': user_roles,
        'maquina_count': maquina_count,
        'faena_count': faena_count,
        'trabajo_count': trabajo_count,
        'faenas_activas': faenas_activas,
        'trabajos_por_estado': trabajos_por_estado,
        'etiquetas_meses': etiquetas_meses,
        'datos_trabajos': datos_trabajos,
        'etiquetas_maquinas': etiquetas_maquinas,
        'datos_maquinas': datos_maquinas
    }
    
    return render(request, 'registros/detalles_empresa.html', context)


    empresa = get_object_or_404(Empresa, pk=pk)

    # Contar máquinas
    maquina_count = Maquina.objects.filter(empresa=empresa).count()
    
    # Contar faenas
    faena_count = Faena.objects.filter(empresa=empresa).count()
    
    # Obtener faenas activas
    faenas_activas = Faena.objects.filter(
        empresa=empresa, 
        estado='activa'
    ).order_by('-fecha_inicio')[:5]  # Mostrar las 5 más recientes
    
    # Contar trabajos
    trabajo_count = Trabajo.objects.filter(empresa=empresa).count()
    
    # Obtener distribución de trabajos por estado
    trabajos_por_estado = {}
    estados_trabajo = Trabajo.objects.filter(empresa=empresa).values('estado').annotate(
        cantidad=Count('id')
    )
    for estado in estados_trabajo:
        trabajos_por_estado[estado['estado']] = estado['cantidad']
    
    # Obtener distribución de máquinas por estado
    maquinas_por_estado = {}
    # Nota: Si tu modelo de Maquina ya no tiene el campo 'estado', este código necesita ser adaptado
    # Por ejemplo, podrías usar un campo diferente o una lógica diferente para categorizar las máquinas
    
    # Opción 1: Si tienes un campo alternativo para representar el estado
    if hasattr(Maquina, 'estado_actual'):  # Reemplaza con el nombre real del campo si existe
        estados_maquina = Maquina.objects.filter(empresa=empresa).values('estado_actual').annotate(
            cantidad=Count('id')
        )
        for estado in estados_maquina:
            maquinas_por_estado[estado['estado_actual']] = estado['cantidad']
    
    # Opción 2: Si no hay un campo de estado, podrías crear categorías basadas en otros atributos
    # Por ejemplo, basado en la fecha de mantenimiento o alguna otra propiedad
    else:
        # Por defecto al menos mostramos máquinas activas
        activas = Maquina.objects.filter(empresa=empresa, activa=True).count()
        inactivas = Maquina.objects.filter(empresa=empresa, activa=False).count()
        
        if activas > 0:
            maquinas_por_estado['operativa'] = activas
        if inactivas > 0:
            maquinas_por_estado['baja'] = inactivas
    
    # Obtener distribución de usuarios por rol
    # Mejorado para asegurar que obtengamos correctamente los recuentos
    user_roles = []
    
    # Obtener todos los grupos existentes primero
    grupos = Group.objects.all()
    
    for grupo in grupos:
        # Contar usuarios en este grupo para esta empresa
        count = User.objects.filter(
            groups=grupo,
            perfil__empresa=empresa
        ).count()
        
        if count > 0:
            user_roles.append({
                'groups__name': grupo.name,
                'user_count': count
            })
    
    # Añadimos "Sin rol" si hay usuarios que no están en ningún grupo
    sin_rol_count = User.objects.filter(
        perfil__empresa=empresa,
        groups__isnull=True
    ).count()
    
    if sin_rol_count > 0:
        user_roles.append({
            'groups__name': 'Sin Rol Asignado',
            'user_count': sin_rol_count
        })
    
    # Crear el diccionario de contexto COMPLETO
    context = {
        'empresa': empresa,
        'user_roles': user_roles,
        'maquina_count': maquina_count,
        'faena_count': faena_count,
        'trabajo_count': trabajo_count,
        'faenas_activas': faenas_activas,
        'trabajos_por_estado': trabajos_por_estado,
        'maquinas_por_estado': maquinas_por_estado,
    }
    
    return render(request, 'registros/detalles_empresa.html', context)
# ---------------------- VISTAS DE USUARIO ----------------------

@login_required
def register_user(request):
    empresa_del_registrador = None

    # --- Determinar la empresa del usuario que registra ---
    try:
        # Permitimos Superuser Y Admin/Supervisor (según tu lógica de permisos)
        # Pero solo asignamos empresa automáticamente si NO es Superuser
        if request.user.is_superuser:
             # Decisión: ¿Permitir a SU registrar sin asignar empresa? ¿O bloquear?
             # Bloquearemos esta vista para SU, deben usar admin u otra interfaz.
             messages.error(request, "Los Superusuarios deben usar la interfaz de administración para crear usuarios y asignar empresas.")
             return redirect('home') # O redirect('admin:index')

        # Para Admin/Supervisor, obtenemos su (única) empresa
        empresa_del_registrador = get_user_empresa(request.user)

        if not empresa_del_registrador:
             messages.error(request, "No se pudo determinar una única empresa asociada a tu cuenta para asignar al nuevo usuario.")
             return redirect('home') # O a donde corresponda

    except MultipleObjectsReturned:
         messages.error(request, "Administras múltiples empresas. No puedes registrar usuarios automáticamente desde aquí.")
         return redirect('home')
    except Exception as e:
         messages.error(request, f"Error al determinar tu empresa: {e}")
         return redirect('home')

    # --- Lógica de Permisos (Ya la tenías, la mantenemos) ---
    # Asegurarse que solo usuarios permitidos accedan (Admin o Supervisor en este caso)
    # Nota: El chequeo de SU ya se hizo arriba para la lógica de la empresa.
    if not request.user.groups.filter(name__in=['Admin', 'Supervisor']).exists():
        # Si llegamos aquí, no es SU y tampoco Admin/Supervisor
        messages.error(request, "No tienes permiso para registrar usuarios.")
        return redirect('home')

    # Si llegamos aquí, empresa_del_registrador tiene la empresa única del Admin/Supervisor

    if request.method == 'POST':
        # Pasamos la empresa determinada al formulario, quitamos user=request.user
        form = UserRegistrationForm(request.POST, empresa=empresa_del_registrador)
        if form.is_valid():
            try:
                # El método save del form ahora maneja la asignación al PerfilUsuario
                user = form.save()
                messages.success(request, f'Usuario {user.username} creado exitosamente en la empresa {empresa_del_registrador.nombre}!')
                 # Ajusta 'listar_usuarios' al nombre de tu URL
                return redirect('listar_usuarios')
            except IntegrityError as e:
                 # Capturar error de username duplicado, etc.
                 if 'UNIQUE constraint failed: auth_user.username' in str(e):
                      messages.error(request, f"El nombre de usuario '{form.cleaned_data.get('username')}' ya existe.")
                 else:
                      messages.error(request, f'Error de base de datos al crear usuario: {str(e)}')
            except ValidationError as e: # Errores del form.clean (ej: max_usuarios)
                 messages.error(request, f"Error de validación: {e.message}")
            except Exception as e:
                 messages.error(request, f'Error inesperado al crear usuario: {str(e)}')
        else:
            # Ajustamos el mensaje de error general
            messages.error(request, 'No se pudo crear el usuario. Por favor, revisa los campos.')
    else:
        # Pasamos la empresa también para GET
        form = UserRegistrationForm(empresa=empresa_del_registrador)

    # La plantilla 'registros/register_user.html' recibe el form SIN el campo 'empresa'
    return render(request, 'registros/register_user.html', {'form': form})


@login_required
def listar_usuarios(request):
    empresa_actual = get_user_empresa(request.user)
    usuarios_query = User.objects.all()

    if not request.user.is_superuser:
        if empresa_actual:
            # Filtrar por perfil__empresa
            usuarios_query = usuarios_query.filter(perfil__empresa=empresa_actual)
        else:
            # Si no es superuser y no tiene empresa, no debería ver usuarios
            usuarios_query = User.objects.none()

    usuarios_list = usuarios_query.select_related('perfil').prefetch_related(
        'groups',
        'perfil__empresa'
    ).order_by('username')

    paginator = Paginator(usuarios_list, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'usuarios': page_obj.object_list,
        'groups': Group.objects.all() # Necesario para el dropdown en la tabla
    }
    return render(request, 'registros/listar_usuarios.html', context)

@login_required
@require_POST # Implementando modal
def eliminar_usuario(request, pk):
    usuario_a_eliminar = get_object_or_404(User, pk=pk)
    empresa_actual = get_user_empresa(request.user)

    permiso_para_eliminar = False
    if request.user.is_superuser:
        permiso_para_eliminar = True
    # Verifica si el usuario a eliminar tiene perfil y pertenece a la misma empresa
    elif hasattr(usuario_a_eliminar, 'perfil') and usuario_a_eliminar.perfil and usuario_a_eliminar.perfil.empresa == empresa_actual:
        # Podrías añadir un check extra: ¿es el usuario logueado un Admin?
        if request.user.groups.filter(name='Admin').exists():
             permiso_para_eliminar = True

    if not permiso_para_eliminar:
        messages.error(request, "No tienes permiso para eliminar este usuario.")
        return redirect('listar_usuarios')

    if usuario_a_eliminar == request.user:
         messages.error(request, "No puedes eliminar tu propia cuenta.")
         return redirect('listar_usuarios')

    try:
        username = usuario_a_eliminar.username
        # Eliminar perfil antes que el usuario
        if hasattr(usuario_a_eliminar, 'perfil'):
            try:
                usuario_a_eliminar.perfil.delete()
            except PerfilUsuario.DoesNotExist:
                pass # No había perfil, no hay problema
        usuario_a_eliminar.delete()
        messages.success(request, f'Usuario "{username}" eliminado exitosamente!')
    except Exception as e:
        messages.error(request, f'Error al eliminar usuario "{usuario_a_eliminar.username}": {str(e)}')

    return redirect('listar_usuarios')

@login_required
@require_POST # Esta vista parece diseñada para procesar datos de la tabla editable
def guardar_cambios_usuarios(request):
    # Esta vista necesita lógica más compleja si se usa para editar en línea
    # Se debe iterar sobre los datos POST para identificar qué usuario y qué campo cambiar
    # Por ahora, solo un mensaje y redirección
    messages.info(request, "Funcionalidad 'Guardar Cambios Múltiples' no implementada completamente en la vista.")
    return redirect('listar_usuarios')


# ---------------------- VISTAS DE MÁQUINA ----------------------

@login_required
def crear_maquina(request):
    empresa_asignada = None
    try:
        # --- Lógica para determinar la empresa automáticamente ---
        if request.user.is_superuser:
            messages.error(request, "Los superusuarios no pueden crear máquinas automáticamente. Utilice la interfaz de administración para especificar la empresa.")
            return redirect('home') # O a la vista de lista o admin

        empresa_asignada = get_user_empresa(request.user)

        if not empresa_asignada:
            messages.error(request, "No se encontró una única empresa asociada a su usuario para la asignación automática.")
            return redirect('home') # O a donde corresponda

    except MultipleObjectsReturned:
        messages.error(request, "Usted administra múltiples empresas. No se puede asignar una automáticamente para crear la máquina.")
        return redirect('home') # O a donde corresponda
    except Exception as e:
        messages.error(request, f"Ocurrió un error al determinar su empresa: {e}")
        return redirect('home')

    # Si llegamos aquí, 'empresa_asignada' contiene la única empresa del usuario

    if request.method == 'POST':
        # Pasamos la empresa determinada al formulario en el constructor
        form = MaquinaForm(request.POST, empresa=empresa_asignada)
        if form.is_valid():
            try:
                # Guardamos sin hacer commit para asignar la empresa manualmente
                maquina = form.save(commit=False)
                maquina.empresa = empresa_asignada # Asignamos la empresa
                maquina.save() # Guardamos la instancia completa en la BD

                # Mensaje de éxito
                messages.success(request, f'Máquina {maquina.nombre} creada exitosamente para la empresa {empresa_asignada.nombre}!')
                
                # Verificar si el usuario quiere crear otra máquina
                if request.POST.get('crear_otra') == '1':
                    return redirect('crear_maquina')
                else:
                    return redirect('listar_maquinas')
                    
            except IntegrityError as e:
                 # Captura errores como unique_together ('nombre', 'empresa')
                if 'UNIQUE constraint failed' in str(e) and 'maquina_nombre_empresa_id_uniq' in str(e): # Ajusta el nombre de la constraint si es diferente
                     messages.error(request, f"Ya existe una máquina con el nombre '{form.cleaned_data.get('nombre')}' en la empresa '{empresa_asignada.nombre}'.")
                else:
                     messages.error(request, f'Error de base de datos al crear máquina: {str(e)}')
            except ValidationError as e: # Errores del form.clean() como max_maquinas
                 messages.error(request, f"Error de validación: {e.message if hasattr(e, 'message') else e}")
            except Exception as e:
                messages.error(request, f'Error inesperado al crear máquina: {str(e)}')
        else:
            # Los errores de campo se mostrarán automáticamente por la plantilla
            messages.error(request, 'Error al crear máquina. Por favor, revise los datos ingresados.')
    else:
        # Para GET, también pasamos la empresa
        form = MaquinaForm(empresa=empresa_asignada)

    # Renderiza la plantilla con el formulario
    return render(request, 'registros/crear_maquina.html', {'form': form})

@login_required
def listar_maquinas(request):
    empresa_actual = get_user_empresa(request.user)
    maquinas_query = Maquina.objects.select_related('empresa')

    if not request.user.is_superuser:
        if empresa_actual:
            maquinas_query = maquinas_query.filter(empresa=empresa_actual)
        else:
            maquinas_query = Maquina.objects.none()

    # Obtiene la lista COMPLETA ordenada
    maquinas_list = maquinas_query.order_by('nombre')

    # Pasa la lista completa con la clave 'maquinas_list' que espera la plantilla
    context = {
        'maquinas_list': maquinas_list,
        'now': timezone.now().date(),  # Agregamos la fecha actual para comparaciones
    }
    return render(request, 'registros/listar_maquinas.html', context)


# --- Vista Editar Maquina (Maneja POST del modal) ---
@login_required
def editar_maquina(request, pk):
    maquina = get_object_or_404(Maquina, pk=pk)
    empresa_actual = get_user_empresa(request.user)

    permiso_para_editar = False
    if request.user.is_superuser:
         permiso_para_editar = True
    elif maquina.empresa == empresa_actual:
         # Asume Admin o Supervisor pueden editar máquinas de su empresa
         if request.user.groups.filter(name__in=['Admin', 'Supervisor']).exists():
              permiso_para_editar = True

    if not permiso_para_editar:
        messages.error(request,"No tienes permiso para editar esta máquina.")
        return redirect('listar_maquinas')

    if request.method == 'POST':
        # Usa el MaquinaForm normal, pasándole el usuario si el form lo necesita
        form = MaquinaForm(request.POST, instance=maquina, empresa=empresa_actual)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, f'Máquina "{maquina.nombre}" actualizada exitosamente!')
            except Exception as e:
                 messages.error(request, f'Error al actualizar máquina: {str(e)}')
        else:
             # Captura errores de validación
             error_list = []
             for field, errors in form.errors.items():
                 error_list.append(f'{field}: {", ".join(errors)}')
             messages.error(request, f"No se pudo actualizar la máquina '{maquina.nombre}'. Errores: {'; '.join(error_list)}")

        # Siempre redirige a la lista después de POST
        return redirect('listar_maquinas')
    else:
         # GET request no esperado si se usa modal
        messages.info(request, "Usa el botón 'Editar' en la lista para modificar máquinas.")
        return redirect('listar_maquinas')


@login_required
@require_POST # Implementando modal
def eliminar_maquina(request, pk):
    maquina = get_object_or_404(Maquina, pk=pk)
    empresa_actual = get_user_empresa(request.user)

    permiso_para_eliminar = False
    if request.user.is_superuser:
        permiso_para_eliminar = True
    elif maquina.empresa == empresa_actual:
         # Asumimos que si pertenece a la empresa, puede eliminar (o añadir check de Rol Admin)
         if request.user.groups.filter(name__in=['Admin', 'Supervisor']).exists(): # Ejemplo de permiso
              permiso_para_eliminar = True

    if not permiso_para_eliminar:
        messages.error(request, "No tienes permiso para eliminar esta máquina.")
        return redirect('listar_maquinas')

    try:
        maquina_nombre = maquina.nombre
        maquina.delete()
        messages.success(request, f'Máquina "{maquina_nombre}" eliminada exitosamente!')
    except Exception as e:
        messages.error(request, f'Error al eliminar máquina: {str(e)}')

    return redirect('listar_maquinas')


# ---------------------- VISTAS DE FAENA ----------------------

@login_required
def crear_faena(request):
    empresa_asignada = None
    try:
        # --- Determinar la empresa del usuario que crea la faena ---
        if request.user.is_superuser:
             messages.error(request, "Los Superusuarios deben usar la interfaz de administración para crear faenas y asignar empresas.")
             return redirect('home') # O admin

        empresa_asignada = get_user_empresa(request.user)

        if not empresa_asignada:
             messages.error(request, "No se pudo determinar una única empresa asociada a tu cuenta para crear la faena.")
             return redirect('home') # O a listar_faenas, etc.

    except MultipleObjectsReturned:
        messages.error(request, "Administras múltiples empresas. No puedes crear faenas automáticamente desde aquí.")
        return redirect('home')
    except Exception as e:
        messages.error(request, f"Error al determinar tu empresa: {e}")
        return redirect('home')

    # Si llegamos aquí, empresa_asignada contiene la empresa única del usuario

    if request.method == 'POST':
        # Pasamos la empresa determinada al formulario
        form = FaenaForm(request.POST, empresa=empresa_asignada)
        if form.is_valid():
            try:
                # Guardamos sin commit para asignar la empresa manualmente
                faena = form.save(commit=False)
                faena.empresa = empresa_asignada # Asignamos la empresa
                faena.save() # Guardamos la instancia completa
                # form.save_m2m() # Si hubiera campos M2M

                # Mensaje de éxito
                messages.success(request, f'Faena {faena.nombre} creada exitosamente para la empresa {empresa_asignada.nombre}!')
                
                # Verificar si el usuario quiere crear otra faena
                if request.POST.get('crear_otra') == '1':
                    return redirect('crear_faena')
                else:
                    return redirect('listar_faenas')
                    
            except IntegrityError as e:
                 # Podría haber unique constraints en Faena (ej: nombre+empresa)
                 messages.error(request, f'Error de base de datos al crear faena: {str(e)}')
            except ValidationError as e: # Errores del form.clean (ej: max_faenas, responsable inválido)
                 # Los errores de campo (como el de responsable) se manejan en el form.errors
                 # Mostramos errores no ligados a campos específicos (non_field_errors)
                 if hasattr(e, 'message_dict'): # Errores de campo específico
                      for field, errors in e.message_dict.items():
                           messages.error(request, f"Error en '{form.fields[field].label}': {'; '.join(errors)}")
                 elif hasattr(e, 'message'): # Errores generales (non_field_errors)
                      messages.error(request, f"Error de validación: {e.message}")
                 else:
                     messages.error(request, f"Error de validación: {e}")

            except Exception as e:
                messages.error(request, f'Error inesperado al crear faena: {str(e)}')
        else:
             messages.error(request, 'Error al crear faena. Por favor, revise los datos ingresados.')
             # Los errores específicos de cada campo se mostrarán en la plantilla
    else:
        # Pasamos la empresa también para GET (necesario para filtrar 'responsable')
        form = FaenaForm(empresa=empresa_asignada)

    # Renderiza la plantilla con el formulario
    return render(request, 'registros/crear_faena.html', {'form': form})

@login_required
def listar_faenas(request):
    empresa_actual = get_user_empresa(request.user)
    # Usamos select_related para optimizar acceso a empresa y responsable
    faenas_query = Faena.objects.select_related('empresa', 'responsable__perfil')

    if not request.user.is_superuser:
        if empresa_actual:
            faenas_query = faenas_query.filter(empresa=empresa_actual)
        else:
            faenas_query = Faena.objects.none()

    # Obtener lista completa ordenada
    faenas_list = faenas_query.order_by('nombre')

    # Obtener lista de supervisores para el modal de edición
    supervisores_list = User.objects.none()
    if request.user.is_superuser:
         supervisores_list = User.objects.filter(groups__name='Supervisor').order_by('username')
    elif empresa_actual:
          supervisores_list = User.objects.filter(
              groups__name='Supervisor',
              perfil__empresa=empresa_actual # Asume relación PerfilUsuario
              ).order_by('username')

    context = {
        'faenas_list': faenas_list,
        'supervisores_list': supervisores_list, # Necesario para el modal
    }
    return render(request, 'registros/listar_faenas.html', context)

# --- Vista Editar Faena (Maneja POST del modal) ---
@login_required
def editar_faena(request, pk):
    faena = get_object_or_404(Faena, pk=pk)
    empresa_actual = get_user_empresa(request.user)

    permiso_para_editar = False
    if request.user.is_superuser:
         permiso_para_editar = True
    elif faena.empresa == empresa_actual:
         # Asume Admin o Supervisor pueden editar faenas de su empresa
         if request.user.groups.filter(name__in=['Admin', 'Supervisor']).exists():
              permiso_para_editar = True

    if not permiso_para_editar:
        messages.error(request,"No tienes permiso para editar esta faena.")
        return redirect('listar_faenas')

    if request.method == 'POST':
        # Usa el FaenaForm normal, pasándole el usuario para filtrar el 'responsable' si es necesario
        form = FaenaForm(request.POST, instance=faena, empresa=empresa_actual)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, f'Faena "{faena.nombre}" actualizada exitosamente!')
            except Exception as e:
                messages.error(request, f'Error al actualizar faena: {str(e)}')
                # Considera mostrar errores del form aquí si es necesario
                # for field, errors in form.errors.items(): ... etc
        else:
             # Captura errores de validación
             error_list = []
             for field, errors in form.errors.items():
                 error_list.append(f'{field}: {", ".join(errors)}')
             messages.error(request, f"No se pudo actualizar la faena '{faena.nombre}'. Errores: {'; '.join(error_list)}")

        # Siempre redirige a la lista después de POST (éxito o fracaso)
        return redirect('listar_faenas')
    else:
        # GET request a esta URL no es esperado si se usa modal
        messages.info(request, "Usa el botón 'Editar' en la lista para modificar faenas.")
        return redirect('listar_faenas')

@login_required
@require_POST # Implementando modal
def eliminar_faena(request, pk):
    faena = get_object_or_404(Faena, pk=pk)
    empresa_actual = get_user_empresa(request.user)

    permiso_para_eliminar = False
    if request.user.is_superuser:
        permiso_para_eliminar = True
    elif faena.empresa == empresa_actual:
         if request.user.groups.filter(name__in=['Admin', 'Supervisor']).exists(): # Ejemplo
              permiso_para_eliminar = True

    if not permiso_para_eliminar:
        messages.error(request, "No tienes permiso para eliminar esta faena.")
        return redirect('listar_faenas')

    try:
        faena_nombre = faena.nombre
        faena.delete()
        messages.success(request, f'Faena "{faena_nombre}" eliminada exitosamente!')
    except Exception as e:
        messages.error(request, f'Error al eliminar faena: {str(e)}')

    return redirect('listar_faenas')


# ---------------------- VISTAS DE TRABAJO ----------------------

@login_required
def crear_trabajo(request):
    empresa_actual = get_user_empresa(request.user)
    puede_crear = request.user.is_superuser or (empresa_actual is not None)

    # Lógica de permisos por rol
    if not request.user.is_superuser and not request.user.groups.filter(name__in=['Admin', 'Supervisor', 'Trabajador']).exists():
         puede_crear = False # Si no tiene rol adecuado

    if not puede_crear:
        messages.error(request,"No tienes permiso o empresa asignada para crear trabajos.")
        return redirect('home')

    if request.method == 'POST':
        form = TrabajoForm(request.POST, user=request.user)
        
        # Si el usuario es trabajador, forzar que el trabajador sea él mismo
        if request.user.groups.filter(name='Trabajador').exists():
            form.data = form.data.copy()  # Hacemos una copia mutable
            form.data['trabajador'] = request.user.id
        
        if form.is_valid():
            try:
                trabajo = form.save(commit=False)
                trabajo.creado_por = request.user
                
                # Si el usuario es trabajador, asegurar que el trabajador sea él mismo
                if request.user.groups.filter(name='Trabajador').exists():
                    trabajo.trabajador = request.user
                
                # Asignar empresa basado en la faena/maquina seleccionada si no se asigna de otra forma
                if not trabajo.empresa:
                     if trabajo.faena:
                          trabajo.empresa = trabajo.faena.empresa
                     elif trabajo.maquina:
                          trabajo.empresa = trabajo.maquina.empresa
                     else: # O asignar la del usuario si es la única forma
                          trabajo.empresa = empresa_actual

                trabajo.save()
                
                # Actualizar el horometro de la máquina si está configurado para hacerlo
                if hasattr(trabajo.maquina, 'horometro_actual') and trabajo.horometro_final:
                    try:
                        maquina = trabajo.maquina
                        maquina.horometro_actual = trabajo.horometro_final
                        maquina.save(update_fields=['horometro_actual'])
                    except Exception as e:
                        # Registrar el error pero no interrumpir el flujo
                        print(f"Error al actualizar horómetro de máquina: {e}")
                
                messages.success(request, '¡Trabajo registrado exitosamente!')
                
                # Si el usuario tiene rol de Trabajador, redirigir a crear otro trabajo
                if 'crear_otra' in request.POST or request.user.groups.filter(name='Trabajador').exists():
                    # Limpiar sesión de formulario
                    if 'trabajo_form_data' in request.session:
                        del request.session['trabajo_form_data']
                    return redirect('crear_trabajo')
                else:
                    return redirect('historial')
            except IntegrityError as e:
                messages.error(request, f'Error de base de datos al crear trabajo: {str(e)}')
            except Exception as e:
                messages.error(request, f'Error inesperado al crear trabajo: {str(e)}')
        else:
            messages.error(request, "Error al crear trabajo. Por favor revisa los campos y corrige los errores.")
    else:
        # Pasa el usuario para filtrar dropdowns
        form = TrabajoForm(user=request.user, initial={'fecha': timezone.now().date()})

    return render(request, 'registros/crear_trabajo.html', {
        'form': form,
        'user': request.user
    })

@login_required
def historial(request):
    # Obtener todos los trabajos con sus relaciones
    trabajos_query = Trabajo.objects.select_related(
        'empresa', 'faena', 'maquina', 'supervisor', 'trabajador'
    )

    # ---- Filtrado por empresa y rol ----
    if not request.user.is_superuser:
        empresa_actual = get_user_empresa(request.user)

        # Primero, limitar a la empresa correspondiente
        if empresa_actual:
            trabajos_query = trabajos_query.filter(empresa=empresa_actual)
        else:
            # Sin empresa asociada ⇒ no ve nada
            trabajos_query = Trabajo.objects.none()

        # Segundo, limitar según el rol del usuario
        if request.user.groups.filter(name='Trabajador').exists():
            trabajos_query = trabajos_query.filter(trabajador=request.user)
        elif request.user.groups.filter(name='Supervisor').exists():
            trabajos_query = trabajos_query.filter(supervisor=request.user)
        # Los Admin de la empresa ven todo lo de la empresa (sin filtro extra)
    else:
        # Superusuario: sin restricciones
        empresa_actual = None

    # Aplicamos filtros manualmente basados en los parámetros GET
    # FECHA: corregimos para asegurar que funcione correctamente
    if 'fecha' in request.GET and request.GET['fecha']:
        fecha_filtro = request.GET['fecha']
        trabajos_query = trabajos_query.filter(fecha=fecha_filtro)
    
    # FAENA: Filtramos por el ID de la faena seleccionada
    if 'faena' in request.GET and request.GET['faena']:
        faena_id = request.GET['faena']
        trabajos_query = trabajos_query.filter(faena_id=faena_id)
    
    # MÁQUINA: Filtramos por el ID de la máquina seleccionada
    if 'maquina' in request.GET and request.GET['maquina']:
        maquina_id = request.GET['maquina']
        trabajos_query = trabajos_query.filter(maquina_id=maquina_id)
    
    # SUPERVISOR: Filtramos por el ID del supervisor seleccionado
    if 'supervisor' in request.GET and request.GET['supervisor']:
        supervisor_id = request.GET['supervisor']
        trabajos_query = trabajos_query.filter(supervisor_id=supervisor_id)
    
    # TRABAJADOR: Filtramos por el ID del trabajador seleccionado
    if 'trabajador' in request.GET and request.GET['trabajador']:
        trabajador_id = request.GET['trabajador']
        trabajos_query = trabajos_query.filter(trabajador_id=trabajador_id)
    
    # ESTADO: Filtramos por el estado seleccionado
    if 'estado' in request.GET and request.GET['estado']:
        estado = request.GET['estado']
        trabajos_query = trabajos_query.filter(estado=estado)

    # ---------------- Dropdowns ----------------
    if request.user.is_superuser:
        # Superusuario: todos
        faenas = Faena.objects.all().order_by('nombre')
        maquinas = Maquina.objects.all().order_by('nombre')
        supervisores = User.objects.filter(groups__name='Supervisor').order_by('username')
        trabajadores = User.objects.filter(groups__name='Trabajador').order_by('username')

    elif request.user.groups.filter(name='Admin').exists():
        # Admin de empresa: todo dentro de su empresa
        faenas = Faena.objects.filter(empresa=empresa_actual).order_by('nombre')
        maquinas = Maquina.objects.filter(empresa=empresa_actual).order_by('nombre')
        supervisores = User.objects.filter(
            groups__name='Supervisor',
            perfil__empresa=empresa_actual
        ).order_by('username')
        trabajadores = User.objects.filter(
            groups__name='Trabajador',
            perfil__empresa=empresa_actual
        ).order_by('username')

    elif request.user.groups.filter(name='Supervisor').exists():
        # Supervisor: solo elementos relacionados a sus propios trabajos
        faena_ids = trabajos_query.values_list('faena_id', flat=True).distinct()
        maquina_ids = trabajos_query.values_list('maquina_id', flat=True).distinct()
        trabajador_ids = trabajos_query.values_list('trabajador_id', flat=True).distinct()

        faenas = Faena.objects.filter(id__in=faena_ids).order_by('nombre')
        maquinas = Maquina.objects.filter(id__in=maquina_ids).order_by('nombre')
        supervisores = User.objects.filter(id=request.user.id)  # Solo él mismo
        trabajadores = User.objects.filter(id__in=trabajador_ids).order_by('username')

    else:
        # Trabajador (u otro rol): elementos propios
        faena_ids = trabajos_query.values_list('faena_id', flat=True).distinct()
        maquina_ids = trabajos_query.values_list('maquina_id', flat=True).distinct()

        faenas = Faena.objects.filter(id__in=faena_ids).order_by('nombre')
        maquinas = Maquina.objects.filter(id__in=maquina_ids).order_by('nombre')
        supervisores = User.objects.filter(groups__name='Supervisor', perfil__empresa=empresa_actual).order_by('username') if empresa_actual else User.objects.none()
        trabajadores = User.objects.filter(id=request.user.id)

    # Ordenar después de filtrar
    trabajos_query = trabajos_query.order_by('-fecha', '-id')

    # Paginación
    paginator = Paginator(trabajos_query, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Añadir todos los datos necesarios al contexto
    context = {
        'filter': {'qs': trabajos_query},  # Mantenemos la estructura esperada por el template
        'page_obj': page_obj,
        'faenas': faenas,
        'maquinas': maquinas,
        'supervisores': supervisores,
        'trabajadores': trabajadores,
        'estados': [
            {'id': 'pendiente', 'nombre': 'Pendiente'},
            {'id': 'aprobado', 'nombre': 'Aprobado'},
            {'id': 'rechazado', 'nombre': 'Rechazado'}
        ]
    }
    
    return render(request, 'registros/historial.html', context)


@login_required
def pendientes(request):
    trabajos_query = Trabajo.objects.select_related(
        'empresa', 'faena', 'maquina', 'supervisor', 'trabajador'
    ).filter(estado='pendiente')

    if not request.user.is_superuser:
        empresa_actual = get_user_empresa(request.user)

        # Limitar a la empresa del usuario
        if empresa_actual:
            trabajos_query = trabajos_query.filter(empresa=empresa_actual)
        else:
            trabajos_query = Trabajo.objects.none()

        # Rol específico
        if request.user.groups.filter(name='Trabajador').exists():
            # Normalmente un trabajador no aprueba pendientes, pero restringimos igual
            trabajos_query = trabajos_query.filter(trabajador=request.user)
        elif request.user.groups.filter(name='Supervisor').exists():
            trabajos_query = trabajos_query.filter(supervisor=request.user)
        # Admin ve todos los pendientes de su empresa

    trabajos_ordenados = trabajos_query.order_by('fecha', 'id') # Ejemplo

    paginator = Paginator(trabajos_ordenados, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'trabajos': page_obj.object_list # Pasar la lista para iterar en template
    }
    return render(request, 'registros/pendientes.html', context)


@login_required
@require_POST
def aprobar_trabajo(request, pk):
    trabajo = get_object_or_404(Trabajo, pk=pk)
    empresa_actual = get_user_empresa(request.user)

    permiso_para_aprobar = False
    if request.user.is_superuser:
         permiso_para_aprobar = True
    # Solo Admin o el Supervisor asignado de la misma empresa pueden aprobar
    elif trabajo.empresa == empresa_actual:
         if request.user.groups.filter(name='Admin').exists() or request.user == trabajo.supervisor:
              permiso_para_aprobar = True

    if not permiso_para_aprobar:
        messages.error(request,"No tienes permiso para aprobar este trabajo")
        return redirect('pendientes')

    if trabajo.estado != 'pendiente':
         messages.warning(request, f"El trabajo ya está en estado '{trabajo.get_estado_display()}'.")
         return redirect('pendientes')

    try:
        # Información General
        if request.POST.get('fecha'):
            trabajo.fecha = request.POST.get('fecha')
        if request.POST.get('faena'):
            trabajo.faena_id = request.POST.get('faena')
        if request.POST.get('maquina'):
            trabajo.maquina_id = request.POST.get('maquina')
        if request.POST.get('trabajo'):
            trabajo.trabajo = request.POST.get('trabajo')

        # Mediciones
        if request.POST.get('tipo_medida'):
            trabajo.tipo_medida = request.POST.get('tipo_medida')
        if request.POST.get('horometro_inicial'):
            trabajo.horometro_inicial = request.POST.get('horometro_inicial')
        if request.POST.get('horometro_final'):
            trabajo.horometro_final = request.POST.get('horometro_final')

        # Recursos
        if request.POST.get('petroleo_litros'):
            trabajo.petroleo_litros = request.POST.get('petroleo_litros')
        if request.POST.get('aceite_tipo'):
            trabajo.aceite_tipo = request.POST.get('aceite_tipo')
        if request.POST.get('aceite_litros'):
            trabajo.aceite_litros = request.POST.get('aceite_litros')

        # Personal
        if request.POST.get('supervisor'):
            trabajo.supervisor_id = request.POST.get('supervisor')
        if request.POST.get('trabajador'):
            trabajo.trabajador_id = request.POST.get('trabajador')

        # Observaciones
        if request.POST.get('observaciones'):
            trabajo.observaciones = request.POST.get('observaciones')

        # Calcular el total de horas
        if trabajo.horometro_final and trabajo.horometro_inicial:
            trabajo.total_horas = float(trabajo.horometro_final) - float(trabajo.horometro_inicial)
        
        # Cambiar el estado a aprobado
        trabajo.estado = 'aprobado'
        trabajo.save()
        
        messages.success(request, 'Trabajo aprobado exitosamente!')
    except Exception as e:
        messages.error(request, f'Error al aprobar trabajo: {str(e)}')

    return redirect('pendientes')


# ---------------------- EXPORTACIÓN Y REPORTES ----------------------

@login_required
def export_historial_xlsx(request):
    trabajos_query = Trabajo.objects.select_related(
        'empresa', 'faena', 'maquina', 'supervisor', 'trabajador'
    )

    if not request.user.is_superuser:
        empresa_actual = get_user_empresa(request.user)

        if empresa_actual:
            trabajos_query = trabajos_query.filter(empresa=empresa_actual)
        else:
            trabajos_query = Trabajo.objects.none()

        # Filtro extra por rol
        if request.user.groups.filter(name='Trabajador').exists():
            trabajos_query = trabajos_query.filter(trabajador=request.user)
        elif request.user.groups.filter(name='Supervisor').exists():
            trabajos_query = trabajos_query.filter(supervisor=request.user)

    trabajo_filter = TrabajoFilter(request.GET, queryset=trabajos_query)
    trabajos_filtrados = trabajo_filter.qs.order_by('-fecha', '-id') # Mismo orden que historial

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    empresa_actual = get_user_empresa(request.user) # Obtener de nuevo por si acaso
    filename_base = "Historial_Trabajos"
    if empresa_actual:
         filename = f"{filename_base}_{empresa_actual.nombre.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}"
    else:
         filename = f"{filename_base}_{timezone.now().strftime('%Y%m%d')}"
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Historial Trabajos'

    columns = [
        'Empresa', 'Fecha', 'Faena', 'Máquina', 'Trabajo', 'Tipo Medida',
        'H/K Inicial', 'H/K Final', 'Total H/K',
        'Petróleo (lts)', 'Aceite (tipo)', 'Aceite (lts)', 'Observaciones',
        'Supervisor', 'Trabajador', 'Estado'
    ]

    header_font = openpyxl.styles.Font(bold=True)
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font

    row_num = 2
    for trabajo in trabajos_filtrados:
        # Formatear valores para Excel
        fecha_str = trabajo.fecha.strftime("%d/%m/%Y") if trabajo.fecha else ''
        hk_inicial = float(trabajo.horometro_inicial) if trabajo.horometro_inicial is not None else ''
        hk_final = float(trabajo.horometro_final) if trabajo.horometro_final is not None else ''
        total_hk = float(trabajo.total_horas) if trabajo.total_horas is not None else ''
        petroleo = float(trabajo.petroleo_litros) if trabajo.petroleo_litros is not None else ''
        aceite_lts = float(trabajo.aceite_litros) if trabajo.aceite_litros is not None else ''

        worksheet.append([
            trabajo.empresa.nombre if trabajo.empresa else '',
            fecha_str,
            trabajo.faena.nombre if trabajo.faena else '',
            trabajo.maquina.nombre if trabajo.maquina else '',
            trabajo.trabajo,
            trabajo.tipo_medida,
            hk_inicial,
            hk_final,
            total_hk,
            petroleo,
            trabajo.aceite_tipo,
            aceite_lts,
            trabajo.observaciones,
            trabajo.supervisor.get_full_name() or trabajo.supervisor.username if trabajo.supervisor else '',
            trabajo.trabajador.get_full_name() or trabajo.trabajador.username if trabajo.trabajador else '',
            trabajo.get_estado_display() # Usa el método del modelo si existe
        ])
        row_num += 1 # No es necesario incrementar manualmente con append

    # Ajustar ancho de columnas
    for col_idx in range(1, len(columns) + 1):
         column_letter = get_column_letter(col_idx)
         worksheet.column_dimensions[column_letter].autosize = True

    workbook.save(response)
    return response


@login_required
def generar_pdf_trabajo(request, pk):
    trabajo = get_object_or_404(Trabajo.objects.select_related(
         'empresa', 'faena', 'maquina', 'supervisor', 'trabajador'
         ), pk=pk)
    user_empresa = get_user_empresa(request.user)

    permiso_para_ver = False
    if request.user.is_superuser:
         permiso_para_ver = True
    elif trabajo.empresa == user_empresa:
         permiso_para_ver = True

    if not permiso_para_ver:
        messages.error(request, "No tienes permiso para generar este reporte.")
        return redirect('historial')

    response = HttpResponse(content_type='application/pdf')
    filename = f"Reporte_{trabajo.faena.nombre if trabajo.faena else 'SinFaena'}_{trabajo.fecha.strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    margin = inch

    # Configuración de colores
    primary_color = colors.HexColor('#0d6efd')  # Color primario de Bootstrap
    secondary_color = colors.HexColor('#6c757d')  # Color secundario
    accent_color = colors.HexColor('#198754')  # Color de éxito
    light_gray = colors.HexColor('#f8f9fa')  # Color de fondo claro
    dark_blue = colors.HexColor('#0a58ca')  # Color azul oscuro para gradiente

    # Encabezado con gradiente
    p.setFillColor(primary_color)
    p.rect(0, height - 2.5*inch, width, 2.5*inch, fill=True, stroke=False)
    
    # Efecto de gradiente (simulado con rectángulos semitransparentes)
    p.setFillColor(dark_blue)
    p.rect(0, height - 2.5*inch, width, 0.5*inch, fill=True, stroke=False)
    
    # Logo y nombre de la empresa
    logo_path_temp = None
    if trabajo.empresa and trabajo.empresa.logo:
        try:
            logo_file = default_storage.open(trabajo.empresa.logo.name)
            fd, logo_path_temp = tempfile.mkstemp(suffix=".png")
            with os.fdopen(fd, 'wb') as temp_logo:
                 temp_logo.write(logo_file.read())
            logo_file.close()
            
            # Dibujar logo con fondo blanco y sombra
            logo_width = 1.5*inch
            logo_height = 1.5*inch
            logo_x = margin
            logo_y = height - margin - logo_height
            
            # Sombra del logo
            p.setFillColor(colors.HexColor('#00000020'))
            p.rect(logo_x + 0.05*inch, logo_y - 0.05*inch, 
                   logo_width + 0.2*inch, logo_height + 0.2*inch, 
                   fill=True, stroke=False)
            
            # Fondo blanco para el logo
            p.setFillColor(colors.white)
            p.rect(logo_x - 0.1*inch, logo_y - 0.1*inch, 
                   logo_width + 0.2*inch, logo_height + 0.2*inch, 
                   fill=True, stroke=False)
            
            # Borde para el logo
            p.setStrokeColor(colors.white)
            p.setLineWidth(2)
            p.rect(logo_x - 0.1*inch, logo_y - 0.1*inch, 
                   logo_width + 0.2*inch, logo_height + 0.2*inch, 
                   fill=False, stroke=True)
            
            # Dibujar el logo
            p.drawImage(logo_path_temp, logo_x, logo_y,
                       width=logo_width, height=logo_height, 
                       preserveAspectRatio=True, mask='auto')
            
            # Nombre de la empresa al lado del logo con sombra de texto
            p.setFont("Helvetica-Bold", 24)
            p.setFillColor(colors.HexColor('#00000040'))
            p.drawString(logo_x + logo_width + 0.32*inch, 
                        height - margin - 0.48*inch, 
                        trabajo.empresa.nombre)
            
            p.setFillColor(colors.white)
            p.drawString(logo_x + logo_width + 0.3*inch, 
                        height - margin - 0.5*inch, 
                        trabajo.empresa.nombre)
            
            # RUT de la empresa debajo del nombre
            p.setFont("Helvetica", 12)
            p.setFillColor(colors.HexColor('#ffffffcc'))  # Blanco semitransparente
            p.drawString(logo_x + logo_width + 0.3*inch, 
                        height - margin - 0.8*inch, 
                        f"RUT: {trabajo.empresa.rut}")
            
        except Exception as e:
            print(f"Error al procesar logo para PDF: {e}")
            # Si hay error con el logo, mostrar solo el nombre de la empresa
            p.setFont("Helvetica-Bold", 24)
            p.setFillColor(colors.white)
            p.drawString(margin, height - margin - 0.5*inch, trabajo.empresa.nombre)
            p.setFont("Helvetica", 12)
            p.setFillColor(colors.HexColor('#ffffffcc'))
            p.drawString(margin, height - margin - 0.8*inch, f"RUT: {trabajo.empresa.rut}")
        finally:
            if logo_path_temp and os.path.exists(logo_path_temp):
                try:
                    os.unlink(logo_path_temp)
                except OSError as e:
                    print(f"Error eliminando archivo temporal de logo: {e}")
    else:
        # Si no hay logo, mostrar solo el nombre de la empresa
        p.setFont("Helvetica-Bold", 24)
        p.setFillColor(colors.white)
        p.drawString(margin, height - margin - 0.5*inch, trabajo.empresa.nombre)
        p.setFont("Helvetica", 12)
        p.setFillColor(colors.HexColor('#ffffffcc'))
        p.drawString(margin, height - margin - 0.8*inch, f"RUT: {trabajo.empresa.rut}")

    # Título del reporte con sombra
    p.setFont("Helvetica-Bold", 20)
    p.setFillColor(colors.HexColor('#00000040'))
    p.drawCentredString(width/2 + 0.02*inch, height - margin - 1.18*inch, "Reporte Diario de Trabajo")
    p.setFillColor(colors.white)
    p.drawCentredString(width/2, height - margin - 1.2*inch, "Reporte Diario de Trabajo")

    # Línea separadora con efecto de gradiente
    p.setStrokeColor(colors.HexColor('#ffffff80'))  # Blanco semitransparente
    p.setLineWidth(2)
    p.line(margin, height - margin - 1.4*inch, width - margin, height - margin - 1.4*inch)
    p.setStrokeColor(colors.white)
    p.setLineWidth(1)
    p.line(margin, height - margin - 1.41*inch, width - margin, height - margin - 1.41*inch)

    # Sección de detalles principales
    y_position = height - margin - 1.8*inch  # Ajustamos el espaciado después del título
    p.setFont("Helvetica-Bold", 14)
    p.setFillColor(primary_color)
    p.drawString(margin, y_position, "Detalles del Registro")
    y_position -= 0.4*inch

    # Crear una tabla con bordes y fondos alternados
    data = [
        ("Fecha:", trabajo.fecha.strftime("%d/%m/%Y") if trabajo.fecha else "N/A"),
        ("Faena:", trabajo.faena.nombre if trabajo.faena else "N/A"),
        ("Máquina:", trabajo.maquina.nombre if trabajo.maquina else "N/A"),
        ("Modelo:", trabajo.maquina.modelo if trabajo.maquina else "N/A"),
        ("Trabajo Realizado:", trabajo.trabajo),
        ("Tipo Medida:", trabajo.tipo_medida),
        (f"{trabajo.tipo_medida} Inicial:", str(trabajo.horometro_inicial)),
        (f"{trabajo.tipo_medida} Final:", str(trabajo.horometro_final)),
        (f"Total {trabajo.tipo_medida}:", str(trabajo.total_horas)),
    ]

    # Dibujar la tabla con estilo
    line_height = 0.3 * inch
    max_label_width = 2.5 * inch
    value_start_x = margin + max_label_width + 0.2*inch
    cell_width = width - 2*margin
    cell_height = line_height

    for i, (label, value) in enumerate(data):
        # Fondo alternado para las filas
        if i % 2 == 0:
            p.setFillColor(light_gray)
            p.rect(margin, y_position - cell_height, cell_width, cell_height, fill=True, stroke=False)
        
        # Bordes de la celda
        p.setStrokeColor(secondary_color)
        p.setLineWidth(0.5)
        p.rect(margin, y_position - cell_height, cell_width, cell_height, fill=False, stroke=True)
        
        # Texto
        p.setFont("Helvetica-Bold", 10)
        p.setFillColor(primary_color)
        p.drawString(margin + 0.1*inch, y_position - 0.2*inch, label)
        
        p.setFont("Helvetica", 10)
        p.setFillColor(colors.black)
        p.drawString(value_start_x, y_position - 0.2*inch, str(value))
        
        y_position -= cell_height

    # Sección de recursos
    y_position -= 0.2*inch
    p.setFont("Helvetica-Bold", 14)
    p.setFillColor(primary_color)
    p.drawString(margin, y_position, "Recursos Utilizados")
    y_position -= 0.4*inch

    recursos_data = [
        ("Petróleo (Lts):", str(trabajo.petroleo_litros) if trabajo.petroleo_litros else "N/A"),
        ("Tipo Aceite:", trabajo.aceite_tipo if trabajo.aceite_tipo else "N/A"),
        ("Aceite (Lts):", str(trabajo.aceite_litros) if trabajo.aceite_litros else "N/A"),
    ]

    for i, (label, value) in enumerate(recursos_data):
        if i % 2 == 0:
            p.setFillColor(light_gray)
            p.rect(margin, y_position - cell_height, cell_width, cell_height, fill=True, stroke=False)
        
        p.setStrokeColor(secondary_color)
        p.setLineWidth(0.5)
        p.rect(margin, y_position - cell_height, cell_width, cell_height, fill=False, stroke=True)
        
        p.setFont("Helvetica-Bold", 10)
        p.setFillColor(primary_color)
        p.drawString(margin + 0.1*inch, y_position - 0.2*inch, label)
        
        p.setFont("Helvetica", 10)
        p.setFillColor(colors.black)
        p.drawString(value_start_x, y_position - 0.2*inch, str(value))
        
        y_position -= cell_height

    # Sección de personal
    y_position -= 0.2*inch
    p.setFont("Helvetica-Bold", 14)
    p.setFillColor(primary_color)
    p.drawString(margin, y_position, "Personal")
    y_position -= 0.4*inch

    personal_data = [
        ("Supervisor:", trabajo.supervisor.get_full_name() if trabajo.supervisor else "N/A"),
        ("Trabajador:", trabajo.trabajador.get_full_name() if trabajo.trabajador else "N/A"),
        ("Estado:", trabajo.get_estado_display()),
    ]

    for i, (label, value) in enumerate(personal_data):
        if i % 2 == 0:
            p.setFillColor(light_gray)
            p.rect(margin, y_position - cell_height, cell_width, cell_height, fill=True, stroke=False)
        
        p.setStrokeColor(secondary_color)
        p.setLineWidth(0.5)
        p.rect(margin, y_position - cell_height, cell_width, cell_height, fill=False, stroke=True)
        
        p.setFont("Helvetica-Bold", 10)
        p.setFillColor(primary_color)
        p.drawString(margin + 0.1*inch, y_position - 0.2*inch, label)
        
        p.setFont("Helvetica", 10)
        p.setFillColor(colors.black)
        p.drawString(value_start_x, y_position - 0.2*inch, str(value))
        
        y_position -= cell_height

    # Observaciones
    y_position -= 0.4*inch
    p.setFont("Helvetica-Bold", 14)
    p.setFillColor(primary_color)
    p.drawString(margin, y_position, "Observaciones")
    y_position -= 0.4*inch

    # Fondo para observaciones con más espacio y margen inferior
    obs_height = 2*inch
    obs_bottom_margin = 1.5*inch  # Aumentamos el margen inferior
    p.setFillColor(light_gray)
    p.rect(margin, y_position - obs_height, cell_width, obs_height, fill=True, stroke=False)
    p.setStrokeColor(secondary_color)
    p.setLineWidth(0.5)
    p.rect(margin, y_position - obs_height, cell_width, obs_height, fill=False, stroke=True)

    # Texto de observaciones
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    styles = getSampleStyleSheet()
    style = ParagraphStyle(
        'CustomStyle',
        parent=styles['Normal'],
        fontSize=10,
        leading=12,
        textColor=colors.black,
        spaceBefore=0.1*inch,
        spaceAfter=0.1*inch,
        leftIndent=0.1*inch,  # Añadimos sangría izquierda
        rightIndent=0.1*inch  # Añadimos sangría derecha
    )
    observaciones_text = trabajo.observaciones if trabajo.observaciones else "Sin observaciones."
    obs_paragraph = Paragraph(observaciones_text, style)
    w_obs, h_obs = obs_paragraph.wrapOn(p, width - 2*margin - 0.2*inch, height)
    obs_paragraph.drawOn(p, margin + 0.1*inch, y_position - h_obs - 0.1*inch)

    # Pie de página con línea separadora y más espacio
    p.setStrokeColor(secondary_color)
    p.setLineWidth(1)
    p.line(margin, margin + 0.8*inch, width - margin, margin + 0.8*inch)  # Subimos la línea
    
    p.setFont("Helvetica", 8)
    p.setFillColor(secondary_color)
    p.drawCentredString(width / 2.0, margin * 0.7,  # Ajustamos la posición del texto
        f"Reporte generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} por {request.user.get_full_name()}")

    p.showPage()
    p.save()
    return response

@login_required
@user_passes_test(lambda u: u.groups.filter(name='Admin').exists())
def dashboard_admin(request):
    # Obtener la empresa del usuario
    empresa = get_user_empresa(request.user)
    if not empresa:
        messages.error(request, "No se encontró una empresa asociada a su usuario.")
        return redirect('home')

    # Conteo de usuarios
    usuarios_count = PerfilUsuario.objects.filter(empresa=empresa).count()
    
    # Conteo de máquinas
    maquina_count = Maquina.objects.filter(empresa=empresa).count()
    
    # Conteo de faenas
    faena_count = Faena.objects.filter(empresa=empresa).count()
    
    # Conteo de trabajos
    trabajo_count = Trabajo.objects.filter(empresa=empresa).count()
    
    # Obtener faenas activas
    faenas_activas = Faena.objects.filter(
        empresa=empresa, 
        estado='activa'
    ).order_by('-fecha_inicio')[:5]
    
    # Obtener trabajos por mes para el gráfico de evolución (últimos 6 meses)
    from django.db.models.functions import TruncMonth
    from datetime import datetime, timedelta
    
    fecha_fin = datetime.now().date() + timedelta(days=1)
    fecha_inicio = (datetime.now() - timedelta(days=180)).date()
    
    trabajos_por_mes = Trabajo.objects.filter(
        empresa=empresa,
        fecha__gte=fecha_inicio,
        fecha__lt=fecha_fin
    ).annotate(
        mes=TruncMonth('fecha')
    ).values('mes').annotate(
        cantidad=Count('id')
    ).order_by('mes')
    
    # Preparar datos para el gráfico
    etiquetas_meses = []
    datos_trabajos = []
    
    meses_map = {}
    current_date = fecha_inicio.replace(day=1)
    while current_date < fecha_fin:
        month_name = current_date.strftime("%b %Y")
        meses_map[month_name] = 0
        etiquetas_meses.append(month_name)
        current_date = (current_date.replace(day=28) + timedelta(days=4)).replace(day=1)
    
    for item in trabajos_por_mes:
        month_name = item['mes'].strftime("%b %Y")
        meses_map[month_name] = item['cantidad']
    
    for month in etiquetas_meses:
        datos_trabajos.append(meses_map.get(month, 0))
    
    # Obtener las máquinas más utilizadas
    maquinas_mas_usadas = Trabajo.objects.filter(
        empresa=empresa
    ).values(
        'maquina__nombre'
    ).annotate(
        cantidad=Count('id')
    ).order_by('-cantidad')[:5]
    
    # Obtener distribución de usuarios por rol
    user_roles = []
    
    # Contar usuarios por rol
    for grupo in Group.objects.all():
        count = User.objects.filter(
            groups=grupo,
            perfil__empresa=empresa
        ).count()
        
        if count > 0:
            user_roles.append({
                'groups__name': grupo.name,
                'user_count': count
            })
    
    # Añadir usuarios sin rol
    sin_rol_count = User.objects.filter(
        perfil__empresa=empresa,
        groups__isnull=True
    ).count()
    
    if sin_rol_count > 0:
        user_roles.append({
            'groups__name': 'Sin Rol Asignado',
            'user_count': sin_rol_count
        })

    # Obtener trabajos pendientes
    trabajos_pendientes = Trabajo.objects.filter(
        empresa=empresa,
        estado='pendiente'
    ).select_related('faena', 'maquina', 'trabajador').order_by('-fecha')[:5]

    # Obtener estadísticas de trabajos por estado
    trabajos_por_estado = Trabajo.objects.filter(
        empresa=empresa
    ).values('estado').annotate(
        cantidad=Count('id')
    )

    # Obtener consumo total de recursos
    consumo_recursos = Trabajo.objects.filter(
        empresa=empresa,
        fecha__gte=fecha_inicio
    ).aggregate(
        total_petroleo=Sum('petroleo_litros'),
        total_aceite=Sum('aceite_litros'),
        total_horas=Sum('total_horas')
    )

    # Obtener trabajos recientes
    trabajos_recientes = Trabajo.objects.filter(
        empresa=empresa
    ).select_related(
        'faena', 'maquina', 'trabajador', 'supervisor'
    ).order_by('-fecha')[:5]
    
    context = {
        'empresa': empresa,
        'usuarios_count': usuarios_count,
        'maquina_count': maquina_count,
        'faena_count': faena_count,
        'trabajo_count': trabajo_count,
        'faenas_activas': faenas_activas,
        'etiquetas_meses': etiquetas_meses,
        'datos_trabajos': datos_trabajos,
        'maquinas_mas_usadas': maquinas_mas_usadas,
        'user_roles': user_roles,
        'trabajos_pendientes': trabajos_pendientes,
        'trabajos_por_estado': trabajos_por_estado,
        'consumo_recursos': consumo_recursos,
        'trabajos_recientes': trabajos_recientes
    }
    
    return render(request, 'registros/dashboard_admin.html', context)

@login_required
def toggle_no_pago(request, empresa_id):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
    from registros.models import Empresa
    empresa = Empresa.objects.get(pk=empresa_id)
    if empresa.no_pago_fecha:
        empresa.no_pago_fecha = None
    else:
        empresa.no_pago_fecha = timezone.now()
    empresa.save()
    return JsonResponse({'success': True, 'no_pago_fecha': empresa.no_pago_fecha})